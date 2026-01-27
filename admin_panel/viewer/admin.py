import openpyxl
import uuid
from datetime import datetime
from django.contrib import admin, messages
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import path
from .models import Activos, Auditorias, Categorias
from .forms import ExcelImportForm

@admin.register(Activos)
class ActivosAdmin(admin.ModelAdmin):
    list_display = ('codigo', 'nombre', 'edificio', 'nivel', 'categoria', 'updated_at')
    search_fields = ('codigo', 'nombre', 'serie')
    list_filter = ('edificio', 'nivel', 'categoria')
    actions = ['export_as_excel']
    change_list_template = "admin/viewer/activos/change_list.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
            path('download-template/', self.download_template),
        ]
        return my_urls + urls

    def export_as_excel(self, request, queryset):
        meta = self.model._meta
        field_names = ['codigo', 'nombre', 'edificio', 'nivel', 'categoria', 'espacio', 'serie']
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(field_names)
        
        for obj in queryset:
            row = []
            for field in field_names:
                val = getattr(obj, field)
                if val is None: val = ""
                row.append(str(val))
            ws.append(row)
            
        wb.save(response)
        return response
    export_as_excel.short_description = "Exportar Seleccionados a Excel"

    def download_template(self, request):
        field_names = ['codigo', 'nombre', 'edificio', 'nivel', 'categoria', 'espacio', 'serie']
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=plantilla_activos.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(field_names)
        
        # Ejemplo
        ws.append(['ACT-001', 'Laptop Dell', 'Edificio A', 'Piso 1', 'Computo', 'Oficina 1', 'SN123456'])
        
        wb.save(response)
        return response

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    
                    count = 0
                    updated = 0
                    errors = 0
                    
                    # Map headers to indices
                    col_map = {}
                    expected = ['codigo', 'nombre', 'edificio', 'nivel', 'categoria', 'espacio', 'serie']
                    
                    # Read header row (1)
                    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if not header_row:
                        messages.error(request, "El archivo está vacío")
                        return redirect("..")

                    for idx, val in enumerate(header_row):
                        if val:
                            clean_val = str(val).lower().strip()
                            if clean_val in expected:
                                col_map[clean_val] = idx
                            
                    if 'codigo' not in col_map or 'nombre' not in col_map:
                         messages.error(request, "El archivo debe tener al menos las columnas 'codigo' y 'nombre'")
                         return redirect("..")

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        try:
                            # Safely get values using col_map
                            def get_val(key):
                                idx = col_map.get(key)
                                if idx is not None and idx < len(row):
                                    v = row[idx]
                                    return str(v).strip() if v is not None else None
                                return None

                            codigo = get_val('codigo')
                            nombre = get_val('nombre')
                            
                            if not codigo or not nombre:
                                continue
                                
                            defaults = {
                                'nombre': nombre,
                                'edificio': get_val('edificio'),
                                'nivel': get_val('nivel'),
                                'categoria': get_val('categoria'),
                                'espacio': get_val('espacio'),
                                'serie': get_val('serie'),
                                'updated_at': datetime.now(),
                                'deleted': 0
                            }
                            
                            # Manual update or create to handle sync_id correctly
                            obj = Activos.objects.filter(codigo=codigo).first()
                            if obj:
                                # Update
                                obj.nombre = nombre
                                obj.edificio = defaults['edificio']
                                obj.nivel = defaults['nivel']
                                obj.categoria = defaults['categoria']
                                obj.espacio = defaults['espacio']
                                obj.serie = defaults['serie']
                                obj.updated_at = defaults['updated_at']
                                obj.deleted = 0
                                obj.save()
                                updated += 1
                            else:
                                # Create
                                Activos.objects.create(
                                    sync_id=str(uuid.uuid4()),
                                    codigo=codigo,
                                    **defaults
                                )
                                count += 1
                                
                        except Exception as e:
                            print(f"Error importing row {row}: {e}")
                            errors += 1
                            
                    messages.success(request, f"Proceso completado. Creados: {count}, Actualizados: {updated}, Errores: {errors}")
                    return redirect("..")
                except Exception as e:
                    messages.error(request, f"Error procesando el archivo: {e}")
                    return redirect("..")
        else:
            form = ExcelImportForm()
            
        return render(
            request, "admin/viewer/activos/import_excel.html", {"form": form}
        )

@admin.register(Auditorias)
class AuditoriasAdmin(admin.ModelAdmin):
    list_display = ('id', 'fecha', 'espacio', 'total_esperados', 'total_escaneados', 'estado', 'updated_at')
    search_fields = ('espacio', 'sync_id')
    list_filter = ('fecha', 'estado')

@admin.register(Categorias)
class CategoriasAdmin(admin.ModelAdmin):
    list_display = ('id', 'nombre', 'parent_id', 'updated_at')
    search_fields = ('nombre',)
